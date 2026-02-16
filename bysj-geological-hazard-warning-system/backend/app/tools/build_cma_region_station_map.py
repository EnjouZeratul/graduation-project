from __future__ import annotations

import argparse
import json
import math
import os
from datetime import datetime, timezone
from typing import Dict, Iterable, List, Tuple

from openpyxl import load_workbook

from app.db import SessionLocal
from app.models import Region


def _norm_header(value: object) -> str:
    return str(value or "").strip().replace(" ", "").replace("\u3000", "")


def _haversine_km(lon1: float, lat1: float, lon2: float, lat2: float) -> float:
    # Earth radius (km)
    r = 6371.0
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return r * c


def _find_header_row(rows: Iterable[Tuple[object, ...]], max_scan: int = 30) -> Tuple[int, Dict[str, int]]:
    """
    Try to locate the header row by searching for '区站号' + '经度' + '纬度'.
    Returns: (row_index_0based, header_map)
    """
    scanned = 0
    for idx, row in enumerate(rows):
        scanned += 1
        headers = [_norm_header(c) for c in row]
        if not headers:
            continue
        # Required columns in your screenshot/doc: 区站号/经度/纬度
        want = {
            "区站号": None,
            "经度": None,
            "纬度": None,
        }
        for col_idx, h in enumerate(headers):
            if not h:
                continue
            if want["区站号"] is None and ("区站号" in h or "Station_Id" in h):
                want["区站号"] = col_idx
            if want["经度"] is None and ("经度" in h or h.lower() in {"lon", "longitude"}):
                want["经度"] = col_idx
            if want["纬度"] is None and ("纬度" in h or h.lower() in {"lat", "latitude"}):
                want["纬度"] = col_idx
        if all(v is not None for v in want.values()):
            return idx, {k: int(v) for k, v in want.items() if v is not None}
        if scanned >= max_scan:
            break
    raise RuntimeError("Failed to locate header row containing 区站号/经度/纬度 in the first sheets.")


def _read_stations_from_xlsx(path: str) -> List[Dict[str, object]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        # Prefer the active sheet; fall back to the first sheet.
        ws = wb.active or wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        header_row_idx, header_map = _find_header_row(rows)
        stations: List[Dict[str, object]] = []
        for row in rows[header_row_idx + 1 :]:
            try:
                station_id = str(row[header_map["区站号"]] or "").strip()
                lon = row[header_map["经度"]]
                lat = row[header_map["纬度"]]
                if not station_id:
                    continue
                if lon is None or lat is None:
                    continue
                lon_f = float(lon)
                lat_f = float(lat)
                # Skip obviously invalid coords
                if not (-180.0 <= lon_f <= 180.0 and -90.0 <= lat_f <= 90.0):
                    continue
                stations.append({"station_id": station_id, "longitude": lon_f, "latitude": lat_f})
            except Exception:
                continue
        if not stations:
            raise RuntimeError("No stations parsed from xlsx (empty result).")
        return stations
    finally:
        wb.close()


def _read_regions_from_db() -> List[Dict[str, object]]:
    db = SessionLocal()
    try:
        rows = db.query(Region).all()
        regions: List[Dict[str, object]] = []
        for r in rows:
            if r.longitude is None or r.latitude is None:
                continue
            regions.append(
                {
                    "code": str(r.code),
                    "name": str(r.name),
                    "longitude": float(r.longitude),
                    "latitude": float(r.latitude),
                }
            )
        return regions
    finally:
        db.close()


def _load_json(path: str) -> Dict[str, str]:
    if not path or not os.path.exists(path):
        return {}
    try:
        raw = open(path, "r", encoding="utf-8").read()
        data = json.loads(raw or "{}")
        if isinstance(data, dict):
            return {str(k): str(v) for k, v in data.items() if str(k).strip() and str(v).strip()}
        return {}
    except Exception:
        return {}


def build_map(
    station_xlsx: str,
    out_path: str,
    *,
    overrides_path: str = "",
    top_k: int = 1,
) -> Dict[str, object]:
    stations = _read_stations_from_xlsx(station_xlsx)
    regions = _read_regions_from_db()
    overrides = _load_json(overrides_path)

    station_points = [(s["station_id"], float(s["longitude"]), float(s["latitude"])) for s in stations]
    mapping: Dict[str, str] = {}

    max_dist = 0.0
    sum_dist = 0.0
    dist_count = 0

    for region in regions:
        code = str(region["code"])
        if code in overrides:
            mapping[code] = overrides[code]
            continue

        lon = float(region["longitude"])
        lat = float(region["latitude"])
        best: List[Tuple[float, str]] = []
        for station_id, slon, slat in station_points:
            d = _haversine_km(lon, lat, slon, slat)
            if not best:
                best = [(d, station_id)]
            else:
                # Keep a small sorted list of top_k.
                best.append((d, station_id))
                best.sort(key=lambda x: x[0])
                if len(best) > top_k:
                    best = best[:top_k]

        if not best:
            continue
        # For now we only output the single nearest station id (string).
        chosen_d, chosen = best[0]
        mapping[code] = chosen
        max_dist = max(max_dist, chosen_d)
        sum_dist += chosen_d
        dist_count += 1

    meta = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "stations_total": len(stations),
        "regions_total": len(regions),
        "mapped_regions": len(mapping),
        "overrides_used": len([k for k in mapping.keys() if k in overrides]),
        "max_distance_km": round(max_dist, 3) if dist_count else None,
        "avg_distance_km": round(sum_dist / max(dist_count, 1), 3) if dist_count else None,
        "station_xlsx": os.path.basename(station_xlsx),
    }

    payload = {"meta": meta, "map": mapping}
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, sort_keys=True)
        f.write("\n")
    return meta


def main() -> None:
    parser = argparse.ArgumentParser(description="Build region_code -> CMA station_id map from China_SURF_Station.xlsx")
    parser.add_argument(
        "--station-xlsx",
        required=True,
        help="Path to China_SURF_Station.xlsx (suggested: /app/app/data/China_SURF_Station.xlsx inside container)",
    )
    parser.add_argument(
        "--out",
        required=True,
        help="Output json path (suggested: /app/app/data/cma_region_station_map.json)",
    )
    parser.add_argument(
        "--overrides",
        default="",
        help="Optional overrides json (region_code -> station_id). Suggested: /app/app/data/cma_region_station_overrides.json",
    )
    args = parser.parse_args()

    meta = build_map(args.station_xlsx, args.out, overrides_path=args.overrides, top_k=1)
    print("OK")
    for k, v in meta.items():
        print(f"{k}: {v}")


if __name__ == "__main__":
    main()

